#!/usr/bin/python3
# -*- coding: utf-8 -*-
"""Module to manage the SQLite and import / export CSV files."""

from datetime import datetime
import logging
import shutil
import sqlite3
from os.path import getmtime
from pathlib import Path

import openpyxl


def check_sqlite(database):
    """Check for SQLite database."""
    dbfile = Path(database)
    if dbfile.is_file():
        logging.info("check_sqlite() == True")
        return True
    else:
        logging.info("check_sqlite() == False")
        return False


def move_sqlite(db_name, db_path, database):
    """Rename SQLite database with timestamp in case it already exists."""
    db_timestamp = str(datetime.fromtimestamp(
        getmtime(database)).strftime('%Y-%m-%d_%H-%M-%S'))
    shutil.move(database, (db_path + db_name + '_' + db_timestamp + '.sqlite'))
    logging.info("move_sqlite() moved %s to %s_%s.sqlite",
                 database, db_name, db_timestamp)


def excel_to_sqlite(excel_file, database, import_sheet, column_names):
    """Actually write the Excel sheet to the SQLite database."""
    conn = sqlite3.connect(database)
    curs = conn.cursor()
    # Create table based on column_names
    curs.execute('CREATE TABLE {} (Id INTEGER PRIMARY KEY, {} TEXT, {} TEXT, {} TEXT, {} TEXT, {} TEXT, {} TEXT, {} TEXT, {} TEXT, {} TEXT)'.
                 format(import_sheet, (*column_names)))
    # Import excel_file into database
    import_wb = openpyxl.load_workbook(excel_file)
    import_ws = import_wb.get_sheet_by_name(import_sheet)
    column_indices = {n: cell.value for n, cell in enumerate(import_ws[1])
                      if cell.value in column_names}
    logging.debug(column_indices)

    for row in import_ws.iter_rows(row_offset=1):
        to_db = list()
        for index, cell in enumerate(row):
            if index in column_indices:
                logging.debug("col: %s, row: %s, content: %s",
                              column_indices[index], index, cell.value)
                to_db.append(cell.value)
        logging.debug(to_db)
        curs.execute('INSERT INTO {} ({}, {}, {}, {}, {}, {}, {}, {}, {}) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)'.
                     format(import_sheet, (*column_names)), to_db)
        logging.info("Inserted values into SQLite.")
        # Delete empty rows
        curs.execute('DELETE FROM {} WHERE {coi} IS NULL'.format(
            import_sheet, coi="Url"))
        logging.info("Deleted empty rows.")
        # Add missing entries in folder column
        curs.execute('UPDATE {} SET {coj} = {coi} WHERE {coj} IS NULL'.format(
            import_sheet, coi="Url", coj="Folder"))
        logging.info("Added missing entries in folder column.")
        # Fix bad entries in folder column (removing all / ~)
        bad_signs = ['https://', 'http://', '/', '\\', '~']
        for sign in bad_signs:
            curs.execute('UPDATE {} SET {coj} = REPLACE({coj}, "{si}", "")'.format(
                import_sheet, coj="Folder", si=sign))
            logging.debug("Removed character '%s' from folder column", sign)
        logging.info("Fixed bad entries in folder column.")
        # Fix bad entries in url column (removing all / ~)
        url_signs = ['https://', 'http://']
        for sign in url_signs:
            curs.execute('UPDATE {} SET {coj} = REPLACE({coj}, "{si}", "")'.format(
                import_sheet, coj="Url", si=sign))
            logging.debug("Removed character '%s' from url column", sign)
        # Trailing / cannot be removed with previous function
        curs.execute('UPDATE {} SET {coj} = REPLACE({coj}, {coj}, rtrim({coj}, "/")) WHERE substr({coj}, -1, 1) LIKE "/"'.format(import_sheet, coj="Url"))
        logging.debug("Removed trailing / from url column")
        logging.info("Fixed bad entries in url column.")

    conn.commit()
    conn.close()


def import_excel(dataset):
    """Import CSV file to SQLite database."""
    sqlite_exists = check_sqlite(dataset['database'])

    if sqlite_exists is True:
        # Move database first before creating a new one.
        move_sqlite(dataset['db_name'], dataset['db_path'],
                    dataset['database'])
        logging.info("import_csv() moved the old database")
        excel_to_sqlite(dataset['excel_file'], dataset['database'],
                        dataset['import_sheet'], dataset['column_names'])

    if sqlite_exists is False:
        logging.info("import_csv() did not move the old database")
        excel_to_sqlite(dataset['excel_file'], dataset['database'],
                        dataset['import_sheet'], dataset['column_names'])


def sqlite_to_excel(excel_file, database, export_sheet, column_names):
    """Actually write the SQLite database to an Excel sheet."""
    conn = sqlite3.connect(database)
    curs = conn.cursor()
    import_wb = openpyxl.load_workbook(excel_file)
    export_ws = import_wb.create_sheet(title=export_sheet)

    # Append rows from database in Excel sheet
    for row in curs.execute('SELECT * FROM import'):
        # TODO: Include header row, remove Id column
        logging.debug(row)
        export_ws.append(row)

    import_wb.save(excel_file)


def export_excel(dataset):
    """Export SQLite database to CSV file."""
    sqlite_exists = check_sqlite(dataset['database'])

    if sqlite_exists is True:
        sqlite_to_excel(dataset['excel_file'], dataset['database'],
                        dataset['export_sheet'], dataset['column_names'])

    if sqlite_exists is False:
        print("No database to export from.")
